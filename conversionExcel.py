import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import time
from openpyxl.styles import PatternFill
from openpyxl import Workbook



def convertirAExcel(horas,frecuencias,hInicial,hMax,hmin,alpha):

    df = pd.DataFrame({
        "Hora_HHMM": [h.strftime("%H:%M") for h in horas],
        "CantidadDeViajes": frecuencias
    })

    plt.figure(figsize=(10, 5))
    plt.plot(df["Hora_HHMM"], df["CantidadDeViajes"], marker='o')
    plt.xticks(rotation=90)
    plt.xlabel("Hora")
    plt.ylabel("Cantidad de Viajes")
    plt.title("Viajes por horario")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("grafico_viajes.png")
    plt.close()

    archivo_excel = "viajes_resultado.xlsx"

    wb = Workbook()
    ws = wb.active

    color_titulo = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
    color_celdas = PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid")

    ws["E2"] = "Parámetros"
    ws["E5"] = f"horaEstablecida: {hInicial}"
    ws["E6"] = f"alpha: {alpha}"
    ws["E7"] = f"horaMaxEsperadaConα : {hMax}"
    ws["E8"] = f"horaMinEsperadaConα : {hmin}"

    for col in ["E", "F", "G", "H"]:
        ws[f"{col}2"].fill = color_titulo
        for row in range(5, 9):
            ws[f"{col}{row}"].fill = color_celdas

    img = ExcelImage("grafico_viajes.png")
    img.anchor = "I2"
    ws.add_image(img)

    wb.save(archivo_excel)

