import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import time
from openpyxl.styles import PatternFill
from openpyxl import Workbook



def convertirAExcel(nombreExcel,cantidad,horas,frecuencias,hInicial,hMax,hmin,alpha,horaCuartoViaje,horaMitad1Viaje,tres_cuartos,vfinal):

    df = pd.DataFrame({
        "Hora_HHMM": [h.strftime("%H:%M") for h in horas],
        "CantidadDeViajes": frecuencias
    })

    plt.figure(figsize=(10, 5))
    plt.plot(df["Hora_HHMM"], df["CantidadDeViajes"], marker='o')
    plt.xticks(rotation=90)
    plt.xlabel("HoraLlegada")
    plt.ylabel("Cantidad de Viajes")
    plt.title("Viajes por horario")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("grafico_viajes.png")
    plt.close()

    archivo_excel = nombreExcel+".xlsx"

    wb = Workbook()
    ws = wb.active

    color_titulo = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
    color_celdas = PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid")

    ws["E2"] = "Parámetros de " +str(cantidad) +" viajes"
    ws["E5"] = f"HoraEstablecida: {hInicial}"
    ws["E6"] = f"Alpha: {alpha}"
    ws["E7"] = f"HoraMinEsperadaConAlpha : {hmin}"
    ws["E8"] = f"HoraMaxEsperadaConAlpha : {hMax}"

    ws["E16"] = f"Parámetros de 1 Solo viaje (Nuevamente Simulado) : "
    ws["E19"] = f"Hora un cuarto del viaje: {horaCuartoViaje}"
    ws["E20"] = f"Hora mitad del Viaje : { horaMitad1Viaje}"
    ws["E21"] = f"Hora tres cuartos del Viaje : {tres_cuartos}"
    ws["E22"] = f"Hora Final del Viaje : {vfinal}"



    for col in ["E", "F", "G", "H"]:
        ws[f"{col}2"].fill = color_titulo
        for row in range(5, 9):
            ws[f"{col}{row}"].fill = color_celdas
    

    
    for col in ["E", "F", "G", "H"]:
        ws[f"{col}16"].fill = color_titulo
        for row in range(19,23):
            ws[f"{col}{row}"].fill = color_celdas

    img = ExcelImage("grafico_viajes.png")
    img.anchor = "I2"
    
    ws.add_image(img)
    wb.save(archivo_excel)

